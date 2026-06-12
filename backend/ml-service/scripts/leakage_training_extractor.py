from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
import time
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

import psycopg2
import psycopg2.extras
import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
GENERATED_ROOT = ROOT / "generated" / "leakage_training"
OUTPUT_ROOT = GENERATED_ROOT / "extracted"
BATCHES_ROOT = GENERATED_ROOT / "batches"
BATCH_INDEX_PATH = GENERATED_ROOT / "batch_index.csv"

EDGE_BASE = "http://localhost:8080"
OUTCOME_BASE = "http://localhost:8081"
INTELLIGENCE_BASE = "http://localhost:8089"

INTENT_DB = {
    "host": "localhost",
    "port": 5436,
    "dbname": "zord_intent_engine_db",
    "user": "intent_user",
    "password": "intent_password",
}

OUTCOME_DB = {
    "host": "localhost",
    "port": 5434,
    "dbname": "zord_outcome_db",
    "user": "outcome_user",
    "password": "outcome_password",
}

INTELLIGENCE_DB = {
    "host": "localhost",
    "port": 5440,
    "dbname": "zord_intelligence",
    "user": "zpi",
    "password": "zpi_secret",
}

SIGNUP_PASSWORD = "LeakageTrain123!"
REQUEST_TIMEOUT = 120
POLL_INTERVAL_SECONDS = 2
POLL_TIMEOUT_SECONDS = 900
DATE_SHIFT_DAYS = 42


def decimal_or_zero(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def fmt_decimal(value: Decimal | None, places: int = 6) -> str:
    if value is None:
        return ""
    quant = Decimal("1") if places == 0 else Decimal("1").scaleb(-places)
    return format(value.quantize(quant), "f")


def normalize_rate(value: Any) -> float | None:
    if value is None:
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if num > 1.0 and num <= 100.0:
        return num / 100.0
    return num


def safe_json_loads(value: Any, default: Any) -> Any:
    if value in (None, "", b""):
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


def percentile_disc(values: list[float], p: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = max(0, math.ceil(p * len(ordered)) - 1)
    return ordered[index]


def parse_isoish_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    text = (value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        pass
    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]
    for pattern in candidates:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def non_empty_rate(values: list[str]) -> float:
    if not values:
        return 0.0
    return sum(1 for value in values if str(value or "").strip()) / len(values)


def dominant_non_empty(values: list[str], fallback: str = "") -> str:
    cleaned = [str(value).strip() for value in values if str(value or "").strip()]
    if not cleaned:
        return fallback
    return Counter(cleaned).most_common(1)[0][0]


def account_beneficiary_key(row: dict[str, str]) -> str:
    parts = [
        row.get("beneficiary_account_number", "").strip(),
        row.get("beneficiary_ifsc", "").strip(),
        row.get("beneficiary_vpa", "").strip(),
        row.get("beneficiary_name", "").strip().lower(),
    ]
    return "|".join(parts)


@dataclass
class BatchSourceData:
    batch_id: str
    anchor_date: date
    intent_rows: list[dict[str, str]]
    truth_manifest_rows: list[dict[str, str]]
    batch_manifest: dict[str, Any]
    runtime_batch_id: str = ""
    runtime_anchor_date: date | None = None
    runtime_intent_path: Path | None = None
    runtime_settlement_path: Path | None = None
    runtime_intent_rows: list[dict[str, str]] | None = None


@dataclass
class BatchSystemState:
    payment_intents: list[dict[str, Any]]
    normalized_rows: list[dict[str, Any]]
    dlq_rows: list[dict[str, Any]]
    observations: list[dict[str, Any]]
    attachment_decisions: list[dict[str, Any]]
    variance_rows: list[dict[str, Any]]
    unresolved_rows: list[dict[str, Any]]
    canonical_batch: dict[str, Any] | None
    batch_attachment_summary: dict[str, Any] | None
    batch_contract: dict[str, Any] | None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def shift_date_text(value: str, days: int, output_has_time: bool) -> str:
    dt = parse_isoish_dt(value)
    if not dt:
        return value
    shifted = dt + timedelta(days=days)
    return shifted.strftime("%Y-%m-%d %H:%M:%S") if output_has_time else shifted.strftime("%Y-%m-%d")


def create_runtime_replay_files(source: BatchSourceData) -> None:
    runtime_dir = OUTPUT_ROOT / "runtime_ingest" / source.runtime_batch_id
    runtime_dir.mkdir(parents=True, exist_ok=True)

    runtime_intent_path = runtime_dir / "intent.csv"
    intent_rows = []
    for row in source.intent_rows:
        updated = dict(row)
        updated["client_batch_ref"] = source.runtime_batch_id
        updated["scheduled_execution_at"] = shift_date_text(updated.get("scheduled_execution_at", ""), DATE_SHIFT_DAYS, True)
        updated["expected_value_date"] = shift_date_text(updated.get("expected_value_date", ""), DATE_SHIFT_DAYS, False)
        intent_rows.append(updated)
    with runtime_intent_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(intent_rows[0].keys()))
        writer.writeheader()
        writer.writerows(intent_rows)

    runtime_settlement_path = runtime_dir / "settlement_replay.xlsx"
    workbook = load_workbook(BATCHES_ROOT / source.batch_id / "settlement_replay.xlsx")
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    for row_idx in range(2, sheet.max_row + 1):
        for header in ("entity_created_at", "payment_captured_at", "dispute_created_at", "settled_at"):
            col_idx = header_map.get(header)
            if not col_idx:
                continue
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value in (None, ""):
                continue
            shifted = shift_date_text(str(cell.value), DATE_SHIFT_DAYS, True)
            cell.value = shifted
    workbook.save(runtime_settlement_path)
    workbook.close()

    source.runtime_anchor_date = source.anchor_date + timedelta(days=DATE_SHIFT_DAYS)
    source.runtime_intent_path = runtime_intent_path
    source.runtime_settlement_path = runtime_settlement_path
    source.runtime_intent_rows = intent_rows


def load_batch_sources() -> list[BatchSourceData]:
    rows = read_csv_rows(BATCH_INDEX_PATH)
    result: list[BatchSourceData] = []
    for row in rows:
        batch_id = row["batch_id"]
        batch_dir = BATCHES_ROOT / batch_id
        batch_manifest = json.loads((batch_dir / "batch_manifest.json").read_text(encoding="utf-8"))
        result.append(
            BatchSourceData(
                batch_id=batch_id,
                anchor_date=datetime.strptime(row["batch_anchor_date"], "%Y-%m-%d").date(),
                intent_rows=read_csv_rows(batch_dir / "intent.csv"),
                truth_manifest_rows=read_csv_rows(batch_dir / "row_truth_manifest.csv"),
                batch_manifest=batch_manifest,
            )
        )
    return sorted(result, key=lambda item: (item.anchor_date, item.batch_id))


class Extractor:
    def __init__(self) -> None:
        self.intent_conn = psycopg2.connect(**INTENT_DB)
        self.outcome_conn = psycopg2.connect(**OUTCOME_DB)
        self.intelligence_conn = psycopg2.connect(**INTELLIGENCE_DB)
        self.intent_conn.autocommit = True
        self.outcome_conn.autocommit = True
        self.intelligence_conn.autocommit = True
        self.session = requests.Session()
        self.tenant_id = ""
        self.access_token = ""
        self.signup_email = ""

    def close(self) -> None:
        self.session.close()
        self.intent_conn.close()
        self.outcome_conn.close()
        self.intelligence_conn.close()

    def signup(self) -> None:
        email = f"leakage.training.{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{uuid.uuid4().hex[:6]}@example.com"
        payload = {
            "tenant_name": f"Leakage Training {datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            "name": "Leakage Trainer",
            "email": email,
            "password": SIGNUP_PASSWORD,
        }
        response = self.session.post(f"{EDGE_BASE}/v1/auth/signup", json=payload, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        body = response.json()
        self.tenant_id = body["user"]["tenant_id"]
        self.access_token = body["access_token"]
        self.signup_email = email

    def login(self, email: str, password: str = SIGNUP_PASSWORD) -> None:
        response = self.session.post(
            f"{EDGE_BASE}/v1/auth/login",
            json={"email": email, "password": password},
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        body = response.json()
        self.tenant_id = body["user"]["tenant_id"]
        self.access_token = body["access_token"]
        self.signup_email = email

    def auth_headers(self, batch_id: str) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self.access_token}",
            "X-Zord-Source-Type": "FILE_UPLOAD",
            "X-Zord-Source-Class": "INTENT",
            "X-Zord-Tenant-Type": "ERP",
            "X-Zord-Source-System": "RAZORPAY",
            "Batch-ID": batch_id,
        }

    def upload_intent_batch(self, source: BatchSourceData) -> None:
        intent_path = source.runtime_intent_path or (BATCHES_ROOT / source.batch_id / "intent.csv")
        with intent_path.open("rb") as handle:
            response = self.session.post(
                f"{EDGE_BASE}/v1/bulk-ingest",
                headers=self.auth_headers(source.runtime_batch_id),
                files={"file": (intent_path.name, handle, "text/csv")},
                timeout=REQUEST_TIMEOUT,
            )
        response.raise_for_status()

    def upload_settlement_batch(self, source: BatchSourceData) -> None:
        replay_path = source.runtime_settlement_path or (BATCHES_ROOT / source.batch_id / "settlement_replay.xlsx")
        with replay_path.open("rb") as handle:
            response = self.session.post(
                f"{OUTCOME_BASE}/v1/settlement/upload",
                params={
                    "tenant_id": self.tenant_id,
                    "psp": "razorpay",
                    "batch_id": source.runtime_batch_id,
                },
                files={"file": (replay_path.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=REQUEST_TIMEOUT,
            )
        response.raise_for_status()

    def run_attachment(self, batch_id: str) -> None:
        payload = {
            "tenant_id": self.tenant_id,
            "job_scope_type": "SETTLEMENT_BATCH",
            "settlement_batch_ref": batch_id,
        }
        response = self.session.post(
            f"{OUTCOME_BASE}/v1/attachment/run",
            json=payload,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()

    def query_all(self, conn: psycopg2.extensions.connection, sql: str, params: tuple[Any, ...]) -> list[dict[str, Any]]:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(row) for row in cur.fetchall()]

    def query_one(self, conn: psycopg2.extensions.connection, sql: str, params: tuple[Any, ...]) -> dict[str, Any] | None:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            row = cur.fetchone()
            return dict(row) if row else None

    def poll_until(self, name: str, condition: Callable[[], bool], timeout_seconds: int = POLL_TIMEOUT_SECONDS) -> None:
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            if condition():
                return
            time.sleep(POLL_INTERVAL_SECONDS)
        raise TimeoutError(f"timed out while waiting for {name}")

    def wait_for_intent_batch(self, source: BatchSourceData) -> None:
        expected_count = len(source.intent_rows)

        def condition() -> bool:
            row = self.query_one(
                self.intent_conn,
                """
                select
                    coalesce((select count(*) from payment_intents where tenant_id = %s and batchid = %s), 0) as pi_count,
                    coalesce((select count(*) from dlq_items where tenant_id = %s and batch_id = %s), 0) as dlq_count
                """,
                (self.tenant_id, source.runtime_batch_id, self.tenant_id, source.runtime_batch_id),
            )
            if not row:
                return False
            return (row["pi_count"] + row["dlq_count"]) >= expected_count

        self.poll_until(f"intent batch {source.batch_id}", condition)

    def intent_batch_ready(self, source: BatchSourceData) -> bool:
        row = self.query_one(
            self.intent_conn,
            """
            select
                coalesce((select count(*) from payment_intents where tenant_id = %s and batchid = %s), 0) as pi_count,
                coalesce((select count(*) from dlq_items where tenant_id = %s and batch_id = %s), 0) as dlq_count
            """,
            (self.tenant_id, source.runtime_batch_id, self.tenant_id, source.runtime_batch_id),
        ) or {"pi_count": 0, "dlq_count": 0}
        return int(row["pi_count"] or 0) + int(row["dlq_count"] or 0) >= len(source.intent_rows)

    def mirror_runtime_batch_into_outcome(self, source: BatchSourceData) -> None:
        rows = self.query_all(
            self.intent_conn,
            """
            select
                intent_id, tenant_id, contract_id, client_payout_ref, amount, currency,
                intended_execution_at, intent_type, provider_hint, proof_readiness_score,
                matchability_score, canonical_hash, governance_state, beneficiary_fingerprint,
                created_at, business_idempotency_key
            from payment_intents
            where tenant_id = %s and batchid = %s
            order by source_row_num nulls last, created_at
            """,
            (self.tenant_id, source.runtime_batch_id),
        )
        if not rows:
            raise RuntimeError(f"no payment_intents available to mirror for {source.runtime_batch_id}")
        tuples = []
        for row in rows:
            tuples.append(
                (
                    row["intent_id"],
                    row["tenant_id"],
                    row["contract_id"],
                    row["client_payout_ref"],
                    source.runtime_batch_id,
                    row.get("business_idempotency_key"),
                    row["amount"],
                    row["currency"],
                    row.get("intended_execution_at"),
                    row.get("intent_type") or "PAYOUT",
                    row.get("provider_hint") or "RAZORPAY",
                    None,
                    row.get("proof_readiness_score") or Decimal("0.80"),
                    row.get("matchability_score") or Decimal("0.75"),
                    row.get("canonical_hash") or "",
                    row.get("governance_state") or "VALID",
                    row.get("beneficiary_fingerprint"),
                    None,
                    row.get("created_at"),
                )
            )
        with self.outcome_conn.cursor() as cur:
            psycopg2.extras.execute_values(
                cur,
                """
                insert into canonical_intents (
                    intent_id, tenant_id, contract_id, client_payout_ref, client_batch_ref,
                    business_idempotency_key, amount, currency_code, intended_execution_at,
                    payout_type, provider_hint, corridor, proof_readiness_score,
                    matchability_score, canonical_hash, governance_state,
                    beneficiary_fingerprint, zord_signature_carrier, created_at
                ) values %s
                on conflict (intent_id) do update set
                    client_batch_ref = excluded.client_batch_ref,
                    amount = excluded.amount,
                    currency_code = excluded.currency_code,
                    intended_execution_at = excluded.intended_execution_at,
                    payout_type = excluded.payout_type,
                    provider_hint = excluded.provider_hint,
                    proof_readiness_score = excluded.proof_readiness_score,
                    matchability_score = excluded.matchability_score,
                    governance_state = excluded.governance_state
                """,
                tuples,
            )
        self.outcome_conn.commit()

    def wait_for_settlement_batch(self, source: BatchSourceData) -> None:
        expected_rows = sum(1 for row in source.truth_manifest_rows if row.get("has_settlement", "").lower() == "true")

        def condition() -> bool:
            row = self.query_one(
                self.outcome_conn,
                """
                select count(*) as obs_count
                from canonical_settlement_observations
                where tenant_id = %s and client_batch_id = %s
                """,
                (self.tenant_id, source.runtime_batch_id),
            )
            return bool(row and row["obs_count"] >= expected_rows)

        self.poll_until(f"settlement batch {source.batch_id}", condition)

    def settlement_batch_ready(self, source: BatchSourceData) -> bool:
        expected_rows = sum(1 for row in source.truth_manifest_rows if row.get("has_settlement", "").lower() == "true")
        row = self.query_one(
            self.outcome_conn,
            """
            select count(*) as obs_count
            from canonical_settlement_observations
            where tenant_id = %s and client_batch_id = %s
            """,
            (self.tenant_id, source.runtime_batch_id),
        ) or {"obs_count": 0}
        return int(row["obs_count"] or 0) >= expected_rows

    def wait_for_attachment_summary(self, source: BatchSourceData) -> None:
        expected_total = len(source.intent_rows)

        def condition() -> bool:
            return self.attachment_batch_ready(source)

        self.poll_until(f"attachment summary {source.batch_id}", condition)

    def attachment_batch_ready(self, source: BatchSourceData) -> bool:
        row = self.query_one(
            self.outcome_conn,
            """
            select total_intent_count, exact_match_count, high_confidence_count,
                   ambiguous_count, unresolved_count, conflicted_count
            from batch_attachment_summaries
            where tenant_id = %s and (batch_id = %s or source_reference = %s)
            order by created_at desc
            limit 1
            """,
            (self.tenant_id, source.runtime_batch_id, source.runtime_batch_id),
        )
        if not row:
            return False
        unresolved_row = self.query_one(
            self.outcome_conn,
            """
            select count(*) as unresolved_count
            from unresolved_intent_records
            where tenant_id = %s and batch_id = %s
            """,
            (self.tenant_id, source.runtime_batch_id),
        ) or {"unresolved_count": 0}
        total_accounted_for = (
            int(row["exact_match_count"] or 0)
            + int(row["high_confidence_count"] or 0)
            + int(row["ambiguous_count"] or 0)
            + int(row["unresolved_count"] or 0)
            + int(row["conflicted_count"] or 0)
            + int(unresolved_row["unresolved_count"] or 0)
        )
        return total_accounted_for >= len(source.intent_rows)

    def wait_for_batch_contracts(self, batch_ids: list[str]) -> None:
        def condition() -> bool:
            row = self.query_one(
                self.intelligence_conn,
                """
                select count(*) as batch_count
                from batch_contracts
                where tenant_id = %s and batch_id = any(%s)
                """,
                (self.tenant_id, batch_ids),
            )
            return bool(row and int(row["batch_count"] or 0) >= len(batch_ids))

        try:
            self.poll_until("batch_contracts", condition, timeout_seconds=90)
        except TimeoutError:
            pass

    def ingest_all(self, batches: list[BatchSourceData]) -> None:
        for batch in batches:
            if not self.intent_batch_ready(batch):
                self.upload_intent_batch(batch)
                self.wait_for_intent_batch(batch)
            self.mirror_runtime_batch_into_outcome(batch)
            if not self.settlement_batch_ready(batch):
                self.upload_settlement_batch(batch)
                self.wait_for_settlement_batch(batch)
            if not self.attachment_batch_ready(batch):
                self.run_attachment(batch.runtime_batch_id)
                self.wait_for_attachment_summary(batch)
        self.wait_for_batch_contracts([batch.runtime_batch_id for batch in batches])

    def fetch_batch_state(self, batch_id: str) -> BatchSystemState:
        payment_intents = self.query_all(
            self.intent_conn,
            """
            select *
            from payment_intents
            where tenant_id = %s and batchid = %s
            order by source_row_num nulls last, created_at
            """,
            (self.tenant_id, batch_id),
        )
        normalized_rows = self.query_all(
            self.intent_conn,
            """
            select nir.*
            from normalized_ingest_records nir
            where nir.tenant_id = %s
              and nir.envelope_id in (
                    select envelope_id from payment_intents where tenant_id = %s and batchid = %s
                    union
                    select envelope_id from dlq_items where tenant_id = %s and batch_id = %s
              )
            order by nir.created_at
            """,
            (self.tenant_id, self.tenant_id, batch_id, self.tenant_id, batch_id),
        )
        dlq_rows = self.query_all(
            self.intent_conn,
            """
            select *
            from dlq_items
            where tenant_id = %s and batch_id = %s
            order by source_row_num nulls last, created_at
            """,
            (self.tenant_id, batch_id),
        )
        observations = self.query_all(
            self.outcome_conn,
            """
            select *
            from canonical_settlement_observations
            where tenant_id = %s and client_batch_id = %s
            order by source_row_ref, created_at
            """,
            (self.tenant_id, batch_id),
        )
        attachment_decisions = self.query_all(
            self.outcome_conn,
            """
            select ad.*
            from attachment_decisions ad
            join canonical_settlement_observations cso
              on cso.settlement_observation_id = ad.settlement_observation_id
            where ad.tenant_id = %s and cso.client_batch_id = %s
            order by ad.created_at
            """,
            (self.tenant_id, batch_id),
        )
        variance_rows = self.query_all(
            self.outcome_conn,
            """
            select vr.*
            from variance_records vr
            join canonical_settlement_observations cso
              on cso.settlement_observation_id = vr.settlement_observation_id
            where vr.tenant_id = %s and cso.client_batch_id = %s
            order by vr.created_at
            """,
            (self.tenant_id, batch_id),
        )
        unresolved_rows = self.query_all(
            self.outcome_conn,
            """
            select *
            from unresolved_intent_records
            where tenant_id = %s and batch_id = %s
            order by created_at
            """,
            (self.tenant_id, batch_id),
        )
        canonical_batch = self.query_one(
            self.intent_conn,
            """
            select *
            from canonical_batches
            where tenant_id = %s and batch_id = %s
            """,
            (self.tenant_id, batch_id),
        )
        batch_attachment_summary = self.query_one(
            self.outcome_conn,
            """
            select *
            from batch_attachment_summaries
            where tenant_id = %s and (batch_id = %s or source_reference = %s)
            order by created_at desc
            limit 1
            """,
            (self.tenant_id, batch_id, batch_id),
        )
        batch_contract = self.query_one(
            self.intelligence_conn,
            """
            select *
            from batch_contracts
            where tenant_id = %s and batch_id = %s
            """,
            (self.tenant_id, batch_id),
        )
        return BatchSystemState(
            payment_intents=payment_intents,
            normalized_rows=normalized_rows,
            dlq_rows=dlq_rows,
            observations=observations,
            attachment_decisions=attachment_decisions,
            variance_rows=variance_rows,
            unresolved_rows=unresolved_rows,
            canonical_batch=canonical_batch,
            batch_attachment_summary=batch_attachment_summary,
            batch_contract=batch_contract,
        )


def compute_raw_batch_features(source: BatchSourceData) -> dict[str, Any]:
    intent_rows = source.runtime_intent_rows or source.intent_rows
    amounts = [decimal_or_zero(row["amount"]) for row in intent_rows]
    amount_floats = [float(amount) for amount in amounts]
    total_amount = sum(amounts, Decimal("0"))
    beneficiary_keys = [account_beneficiary_key(row) for row in intent_rows]
    beneficiary_counts = Counter(beneficiary_keys)
    repeated_amount = sum(
        decimal_or_zero(row["amount"])
        for row, key in zip(intent_rows, beneficiary_keys)
        if beneficiary_counts[key] > 1
    )
    pair_counts = Counter((key, str(decimal_or_zero(row["amount"]))) for row, key in zip(intent_rows, beneficiary_keys))
    schedule_times = [
        parse_isoish_dt(row.get("scheduled_execution_at") or row.get("expected_value_date"))
        for row in intent_rows
    ]
    schedule_times = [value for value in schedule_times if value is not None]
    batch_ts = min(schedule_times) if schedule_times else datetime.combine(source.runtime_anchor_date or source.anchor_date, datetime.min.time())
    return {
        "batch_total_intended_amount_minor": total_amount,
        "batch_intent_count": len(intent_rows),
        "batch_avg_amount_minor": (total_amount / len(amounts)) if amounts else Decimal("0"),
        "batch_max_amount_minor": max(amounts) if amounts else Decimal("0"),
        "batch_min_amount_minor": min(amounts) if amounts else Decimal("0"),
        "batch_amount_stddev": Decimal(str(statistics.pstdev(amount_floats))) if len(amount_floats) > 1 else Decimal("0"),
        "batch_same_beneficiary_amount_density": float(repeated_amount / total_amount) if total_amount else 0.0,
        "batch_max_pair_count": max(pair_counts.values()) if pair_counts else 0,
        "client_payout_ref_coverage_rate": non_empty_rate([row.get("client_payout_ref", "") for row in intent_rows]),
        "currency": dominant_non_empty([row.get("currency", "") for row in intent_rows], fallback="INR"),
        "source_system": dominant_non_empty([row.get("source_system", "") for row in intent_rows], fallback="UNKNOWN"),
        "rail": dominant_non_empty(
            [row.get("rail_hint", "") or row.get("payment_method", "") for row in intent_rows],
            fallback="UNKNOWN",
        ),
        "created_hour": batch_ts.hour,
        "created_day_of_week": batch_ts.weekday(),
        "weekend_flag": 1 if batch_ts.weekday() >= 5 else 0,
        "_business_batch_ts": batch_ts,
    }


def compute_intent_engine_features(state: BatchSystemState) -> dict[str, Any]:
    received_count = len(state.payment_intents) + len(state.dlq_rows)
    mapping_scores = [
        normalize_rate(row.get("mapping_confidence_score"))
        for row in state.payment_intents
        if normalize_rate(row.get("mapping_confidence_score")) is not None
    ]
    schema_scores = [
        normalize_rate(row.get("schema_completeness_score"))
        for row in state.payment_intents
        if normalize_rate(row.get("schema_completeness_score")) is not None
    ]
    intent_types = [str(row.get("intent_type") or "").strip() for row in state.payment_intents]
    unknown_keys: set[str] = set()
    parse_success_flags: list[int] = []
    missing_required_flags: list[int] = []
    for row in state.normalized_rows:
        unmapped = safe_json_loads(row.get("unmapped_json"), {})
        if isinstance(unmapped, dict):
            unknown_keys.update(str(key) for key in unmapped.keys())
        gap_count = int(row.get("required_field_gap_count") or 0)
        parse_success_flags.append(1 if gap_count == 0 else 0)
        missing_required_flags.append(1 if gap_count > 0 else 0)

    invalid_amount_count = sum(1 for row in state.dlq_rows if str(row.get("reason_code") or "").upper() == "INVALID_AMOUNT")
    invalid_beneficiary_count = sum(
        1
        for row in state.dlq_rows
        if str(row.get("reason_code") or "").upper() in {"INVALID_INSTRUMENT", "INVALID_BENEFICIARY"}
        or "beneficiary" in str(row.get("error_detail") or "").lower()
    )
    canonical_batch = state.canonical_batch or {}
    canonicalization_success_rate = normalize_rate(canonical_batch.get("canonicalization_success_rate"))
    if canonicalization_success_rate is None:
        canonicalization_success_rate = (len(state.payment_intents) / received_count) if received_count else 0.0
    parse_success_rate = (
        sum(parse_success_flags) / len(parse_success_flags)
        if parse_success_flags
        else (len(state.payment_intents) / received_count if received_count else 0.0)
    )
    return {
        "intent_type": dominant_non_empty(intent_types, fallback="UNKNOWN"),
        "parse_success_rate": parse_success_rate,
        "mapping_confidence_score": (sum(mapping_scores) / len(mapping_scores)) if mapping_scores else None,
        "required_field_completeness_rate": (sum(schema_scores) / len(schema_scores)) if schema_scores else None,
        "canonicalization_error_rate": max(0.0, 1.0 - float(canonicalization_success_rate or 0.0)),
        "missing_required_field_rate": (sum(missing_required_flags) / len(missing_required_flags)) if missing_required_flags else 0.0,
        "unknown_column_count": len(unknown_keys),
        "invalid_amount_rate": (invalid_amount_count / received_count) if received_count else 0.0,
        "invalid_beneficiary_rate": (invalid_beneficiary_count / received_count) if received_count else 0.0,
    }


def compute_outcome_current_features(source: BatchSourceData, state: BatchSystemState) -> dict[str, Any]:
    provider_keys = [
        str(obs.get("source_system_id") or obs.get("source_system") or "").strip().lower()
        for obs in state.observations
        if str(obs.get("source_system_id") or obs.get("source_system") or "").strip()
    ]
    provider_key = dominant_non_empty(provider_keys, fallback="razorpay")
    confidence_scores = [float(row["confidence_score"]) for row in state.attachment_decisions if row.get("confidence_score") is not None]
    score_margins = [float(row["score_margin"]) for row in state.attachment_decisions if row.get("score_margin") is not None]
    candidate_sizes = [int(row["candidate_set_size"] or 0) for row in state.attachment_decisions]
    carrier_scores = [float(obs["carrier_richness_score"]) for obs in state.observations if obs.get("carrier_richness_score") is not None]
    current_delays = [
        float(row["settlement_delay_days"])
        for row in state.variance_rows
        if row.get("settlement_delay_days") is not None
    ]
    return {
        "provider_key": provider_key,
        "avg_attachment_confidence": (sum(confidence_scores) / len(confidence_scores)) if confidence_scores else None,
        "low_confidence_rate": (
            sum(1 for value in confidence_scores if value < 0.70) / len(confidence_scores)
            if confidence_scores
            else None
        ),
        "candidate_collision_rate": (
            sum(1 for value in candidate_sizes if value > 1) / len(candidate_sizes)
            if candidate_sizes
            else None
        ),
        "avg_score_margin": (sum(score_margins) / len(score_margins)) if score_margins else None,
        "carrier_completeness_rate": (
            sum(1 for value in carrier_scores if value >= 0.60) / len(carrier_scores)
            if carrier_scores
            else None
        ),
        "current_batch_settlement_delay_p50_days": percentile_disc(current_delays, 0.50),
        "current_batch_settlement_delay_p95_days": percentile_disc(current_delays, 0.95),
    }


def compute_label(source: BatchSourceData, state: BatchSystemState, total_intended_amount: Decimal) -> dict[str, Any]:
    batch_contract = state.batch_contract or {}
    contract_unmatched = decimal_or_zero(batch_contract.get("unmatched_amount_minor"))
    contract_under = decimal_or_zero(batch_contract.get("under_settlement_amount_minor"))
    contract_reversal = decimal_or_zero(batch_contract.get("reversal_exposure_minor"))
    contract_total = contract_unmatched + contract_under + contract_reversal

    outcome_unmatched = sum((decimal_or_zero(row.get("amount")) for row in state.unresolved_rows), Decimal("0"))
    outcome_under = sum(
        (abs(decimal_or_zero(row.get("amount_variance"))) for row in state.variance_rows if str(row.get("variance_type") or "") == "UNDER_SETTLEMENT"),
        Decimal("0"),
    )
    outcome_reversal = sum(
        (abs(decimal_or_zero(row.get("amount_variance"))) for row in state.variance_rows if str(row.get("variance_type") or "") == "REVERSAL"),
        Decimal("0"),
    )
    outcome_total = outcome_unmatched + outcome_under + outcome_reversal

    truth_label = decimal_or_zero(source.batch_manifest["truth_label"]["predicted_leakage_rate"])
    truth_leakage_amount = decimal_or_zero(source.batch_manifest["expected_total_leakage_amount"])
    contract_rate = (contract_total / total_intended_amount) if total_intended_amount else Decimal("0")
    outcome_rate = (outcome_total / total_intended_amount) if total_intended_amount else Decimal("0")
    if abs(outcome_total - truth_leakage_amount) < abs(contract_total - truth_leakage_amount):
        best_source = "outcome_engine_fallback"
        best_unmatched = outcome_unmatched
        best_under = outcome_under
        best_reversal = outcome_reversal
        best_total = outcome_total
        best_rate = outcome_rate
    else:
        best_source = "intelligence.batch_contracts"
        best_unmatched = contract_unmatched
        best_under = contract_under
        best_reversal = contract_reversal
        best_total = contract_total
        best_rate = contract_rate
    return {
        "system_contract_unmatched_intent_amount_minor": contract_unmatched,
        "system_contract_under_settlement_amount_minor": contract_under,
        "system_contract_confirmed_reversal_amount_minor": contract_reversal,
        "system_contract_total_leakage_amount_minor": contract_total,
        "system_contract_predicted_leakage_rate": contract_rate,
        "system_outcome_unmatched_intent_amount_minor": outcome_unmatched,
        "system_outcome_under_settlement_amount_minor": outcome_under,
        "system_outcome_confirmed_reversal_amount_minor": outcome_reversal,
        "system_outcome_total_leakage_amount_minor": outcome_total,
        "system_outcome_predicted_leakage_rate": outcome_rate,
        "system_unmatched_intent_amount_minor": best_unmatched,
        "system_under_settlement_amount_minor": best_under,
        "system_confirmed_reversal_amount_minor": best_reversal,
        "system_total_leakage_amount_minor": best_total,
        "system_predicted_leakage_rate": best_rate,
        "best_system_label_source": best_source,
        "predicted_leakage_rate": truth_label,
        "target_leakage_amount_minor": truth_leakage_amount,
        "training_label_source": "truth_manifest",
        "truth_total_leakage_amount_minor": truth_leakage_amount,
        "truth_predicted_leakage_rate": truth_label,
        "system_label_abs_delta_amount_minor": abs(best_total - truth_leakage_amount),
        "system_label_abs_delta_rate": abs(best_rate - truth_label),
    }


def build_history_events(source: BatchSourceData, state: BatchSystemState, provider_key: str) -> dict[str, list[dict[str, Any]]]:
    batch_ts = datetime.combine(source.anchor_date, datetime.min.time())
    observation_events: list[dict[str, Any]] = []
    observation_by_id = {str(obs["settlement_observation_id"]): obs for obs in state.observations}
    for obs in state.observations:
        obs_provider = str(obs.get("source_system_id") or obs.get("source_system") or provider_key).strip().lower() or provider_key
        observation_ts = obs.get("observation_timestamp") or obs.get("created_at")
        observation_events.append(
            {
                "batch_id": source.batch_id,
                "provider_key": obs_provider,
                "business_ts": parse_isoish_dt(str(observation_ts)) if observation_ts else batch_ts,
                "missing_provider_ref": 1 if not str(obs.get("provider_reference") or "").strip() else 0,
                "missing_client_ref": 1 if not str(obs.get("client_reference_candidate") or "").strip() else 0,
            }
        )
    variance_events: list[dict[str, Any]] = []
    for row in state.variance_rows:
        obs = observation_by_id.get(str(row.get("settlement_observation_id")))
        obs_provider = str((obs or {}).get("source_system_id") or (obs or {}).get("source_system") or provider_key).strip().lower() or provider_key
        event_ts = row.get("created_at")
        delay_days = row.get("settlement_delay_days")
        variance_events.append(
            {
                "batch_id": source.batch_id,
                "provider_key": obs_provider,
                "business_ts": parse_isoish_dt(str(event_ts)) if event_ts else batch_ts,
                "delay_days": float(delay_days) if delay_days is not None else None,
                "provider_ref_missing_flag": 1 if row.get("provider_ref_missing_flag") else 0,
            }
        )
    return {"observation_events": observation_events, "variance_events": variance_events}


def compute_historical_features(
    batch_ts: datetime,
    provider_key: str,
    prior_observation_events: list[dict[str, Any]],
    prior_variance_events: list[dict[str, Any]],
) -> dict[str, Any]:
    window_start = batch_ts - timedelta(days=30)

    def in_window(event: dict[str, Any]) -> bool:
        ts = event["business_ts"]
        return ts is not None and window_start <= ts < batch_ts

    provider_observations = [
        event for event in prior_observation_events
        if in_window(event) and event["provider_key"] == provider_key
    ]
    provider_variances = [
        event for event in prior_variance_events
        if in_window(event) and event["provider_key"] == provider_key and event["delay_days"] is not None
    ]
    tenant_variances = [
        event for event in prior_variance_events
        if in_window(event) and event["delay_days"] is not None
    ]

    provider_delays = [event["delay_days"] for event in provider_variances]
    tenant_delays = [event["delay_days"] for event in tenant_variances]
    return {
        "provider_missing_provider_ref_rate": (
            sum(event["missing_provider_ref"] for event in provider_observations) / len(provider_observations)
            if provider_observations
            else None
        ),
        "provider_missing_client_ref_rate": (
            sum(event["missing_client_ref"] for event in provider_observations) / len(provider_observations)
            if provider_observations
            else None
        ),
        "provider_settlement_delay_p50_days": percentile_disc(provider_delays, 0.50),
        "provider_settlement_delay_p95_days": percentile_disc(provider_delays, 0.95),
        "settlement_delay_p50_days": percentile_disc(tenant_delays, 0.50),
        "settlement_delay_p95_days": percentile_disc(tenant_delays, 0.95),
    }


def feature_catalog() -> list[dict[str, Any]]:
    return [
        {"feature": "batch_total_intended_amount_minor", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_intent_count", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_avg_amount_minor", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_max_amount_minor", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_min_amount_minor", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_amount_stddev", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_same_beneficiary_amount_density", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "batch_max_pair_count", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "client_payout_ref_coverage_rate", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "currency", "source": "raw batch intent.csv + payment_intents", "phase": "intent_time", "safe_for_model": True},
        {"feature": "source_system", "source": "raw batch intent.csv + payment_intents", "phase": "intent_time", "safe_for_model": True},
        {"feature": "intent_type", "source": "payment_intents", "phase": "intent_time", "safe_for_model": True},
        {"feature": "created_hour", "source": "raw batch intent.csv scheduled_execution_at", "phase": "intent_time", "safe_for_model": True},
        {"feature": "created_day_of_week", "source": "raw batch intent.csv scheduled_execution_at", "phase": "intent_time", "safe_for_model": True},
        {"feature": "weekend_flag", "source": "raw batch intent.csv scheduled_execution_at", "phase": "intent_time", "safe_for_model": True},
        {"feature": "rail", "source": "raw batch intent.csv", "phase": "intent_time", "safe_for_model": True},
        {"feature": "provider_key", "source": "canonical_settlement_observations source_system_id fallback raw source", "phase": "mostly_intent_time_proxy", "safe_for_model": True},
        {"feature": "parse_success_rate", "source": "normalized_ingest_records", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "mapping_confidence_score", "source": "payment_intents", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "required_field_completeness_rate", "source": "payment_intents schema_completeness_score", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "canonicalization_error_rate", "source": "canonical_batches / dlq_items", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "missing_required_field_rate", "source": "normalized_ingest_records", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "unknown_column_count", "source": "normalized_ingest_records.unmapped_json distinct keys", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "invalid_amount_rate", "source": "dlq_items reason_code", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "invalid_beneficiary_rate", "source": "dlq_items reason_code/error_detail heuristic", "phase": "post_ingest_pre_settlement", "safe_for_model": True},
        {"feature": "provider_missing_provider_ref_rate", "source": "outcome observations rolling 30d", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "provider_missing_client_ref_rate", "source": "outcome observations rolling 30d", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "provider_settlement_delay_p50_days", "source": "variance_records rolling 30d by provider", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "provider_settlement_delay_p95_days", "source": "variance_records rolling 30d by provider", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "avg_attachment_confidence", "source": "attachment_decisions", "phase": "post_settlement", "safe_for_model": False},
        {"feature": "low_confidence_rate", "source": "attachment_decisions", "phase": "post_settlement", "safe_for_model": False},
        {"feature": "candidate_collision_rate", "source": "attachment_decisions", "phase": "post_settlement", "safe_for_model": False},
        {"feature": "avg_score_margin", "source": "attachment_decisions", "phase": "post_settlement", "safe_for_model": False},
        {"feature": "carrier_completeness_rate", "source": "canonical_settlement_observations", "phase": "post_settlement", "safe_for_model": False},
        {"feature": "settlement_delay_p50_days", "source": "variance_records rolling 30d tenant-wide", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "settlement_delay_p95_days", "source": "variance_records rolling 30d tenant-wide", "phase": "historical_pre_settlement", "safe_for_model": True},
        {"feature": "predicted_leakage_rate", "source": "truth manifest ground truth", "phase": "label", "safe_for_model": False},
    ]


def format_row(row: dict[str, Any]) -> dict[str, Any]:
    formatted: dict[str, Any] = {}
    decimal_columns = {
        "batch_total_intended_amount_minor",
        "batch_avg_amount_minor",
        "batch_max_amount_minor",
        "batch_min_amount_minor",
        "batch_amount_stddev",
        "system_contract_unmatched_intent_amount_minor",
        "system_contract_under_settlement_amount_minor",
        "system_contract_confirmed_reversal_amount_minor",
        "system_contract_total_leakage_amount_minor",
        "system_outcome_unmatched_intent_amount_minor",
        "system_outcome_under_settlement_amount_minor",
        "system_outcome_confirmed_reversal_amount_minor",
        "system_outcome_total_leakage_amount_minor",
        "system_unmatched_intent_amount_minor",
        "system_under_settlement_amount_minor",
        "system_confirmed_reversal_amount_minor",
        "system_total_leakage_amount_minor",
        "target_leakage_amount_minor",
        "truth_total_leakage_amount_minor",
        "system_label_abs_delta_amount_minor",
    }
    rate_like_columns = {
        "batch_same_beneficiary_amount_density",
        "client_payout_ref_coverage_rate",
        "parse_success_rate",
        "mapping_confidence_score",
        "required_field_completeness_rate",
        "canonicalization_error_rate",
        "missing_required_field_rate",
        "invalid_amount_rate",
        "invalid_beneficiary_rate",
        "provider_missing_provider_ref_rate",
        "provider_missing_client_ref_rate",
        "avg_attachment_confidence",
        "low_confidence_rate",
        "candidate_collision_rate",
        "carrier_completeness_rate",
        "predicted_leakage_rate",
        "system_predicted_leakage_rate",
        "system_contract_predicted_leakage_rate",
        "system_outcome_predicted_leakage_rate",
        "truth_predicted_leakage_rate",
        "system_label_abs_delta_rate",
    }
    simple_float_columns = {
        "avg_score_margin",
        "provider_settlement_delay_p50_days",
        "provider_settlement_delay_p95_days",
        "settlement_delay_p50_days",
        "settlement_delay_p95_days",
        "current_batch_settlement_delay_p50_days",
        "current_batch_settlement_delay_p95_days",
    }
    for key, value in row.items():
        if key.startswith("_"):
            continue
        if isinstance(value, Decimal):
            if key in decimal_columns or key in rate_like_columns:
                formatted[key] = fmt_decimal(value, places=6)
            else:
                formatted[key] = format(value, "f")
        elif key in rate_like_columns and value is not None:
            formatted[key] = f"{float(value):.6f}"
        elif key in simple_float_columns and value is not None:
            formatted[key] = f"{float(value):.4f}"
        else:
            formatted[key] = value
    return formatted


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"no rows to write for {path}")
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--email", help="resume against an existing tenant session email")
    parser.add_argument("--password", default=SIGNUP_PASSWORD, help="password for --email login")
    args = parser.parse_args()

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    batches = load_batch_sources()
    extractor = Extractor()
    try:
        if args.email:
            extractor.login(args.email, args.password)
        else:
            extractor.signup()
        for batch in batches:
            batch.runtime_batch_id = f"{batch.batch_id}__{extractor.tenant_id.split('-')[0]}"
            create_runtime_replay_files(batch)
        extractor.ingest_all(batches)

        states = {batch.batch_id: extractor.fetch_batch_state(batch.runtime_batch_id) for batch in batches}
        prior_observation_events: list[dict[str, Any]] = []
        prior_variance_events: list[dict[str, Any]] = []
        full_rows: list[dict[str, Any]] = []

        for batch in batches:
            state = states[batch.batch_id]
            raw_features = compute_raw_batch_features(batch)
            intent_features = compute_intent_engine_features(state)
            current_outcome = compute_outcome_current_features(batch, state)
            history_features = compute_historical_features(
                batch_ts=raw_features["_business_batch_ts"],
                provider_key=current_outcome["provider_key"],
                prior_observation_events=prior_observation_events,
                prior_variance_events=prior_variance_events,
            )
            label_features = compute_label(
                batch,
                state,
                total_intended_amount=raw_features["batch_total_intended_amount_minor"],
            )

            row = {
                "tenant_id": extractor.tenant_id,
                "batch_id": batch.batch_id,
                "runtime_batch_id": batch.runtime_batch_id,
                "batch_template": batch.batch_manifest["batch_template"],
                "batch_anchor_date": batch.anchor_date.isoformat(),
                "runtime_anchor_date": (batch.runtime_anchor_date or batch.anchor_date).isoformat(),
                **raw_features,
                **intent_features,
                **current_outcome,
                **history_features,
                **label_features,
                "system_contract_available": 1 if state.batch_contract else 0,
                "payment_intent_row_count": len(state.payment_intents),
                "settlement_observation_row_count": len(state.observations),
                "attachment_decision_row_count": len(state.attachment_decisions),
                "variance_row_count": len(state.variance_rows),
            }
            full_rows.append(format_row(row))
            history_events = build_history_events(batch, state, current_outcome["provider_key"])
            prior_observation_events.extend(history_events["observation_events"])
            prior_variance_events.extend(history_events["variance_events"])

        safe_features = {
            item["feature"]
            for item in feature_catalog()
            if item["safe_for_model"] and item["phase"] != "label"
        }
        safe_rows: list[dict[str, Any]] = []
        for row in full_rows:
            safe_row = {key: value for key, value in row.items() if key in safe_features or key in {
                "tenant_id",
                "batch_id",
                "batch_template",
                "batch_anchor_date",
                "predicted_leakage_rate",
                "system_predicted_leakage_rate",
                "system_total_leakage_amount_minor",
                "target_leakage_amount_minor",
                "truth_total_leakage_amount_minor",
                "truth_predicted_leakage_rate",
                "best_system_label_source",
                "training_label_source",
                "system_label_abs_delta_amount_minor",
                "system_label_abs_delta_rate",
                "system_contract_available",
            }}
            safe_rows.append(safe_row)

        write_csv(OUTPUT_ROOT / "training_dataset_full.csv", full_rows)
        write_csv(OUTPUT_ROOT / "training_dataset_intent_safe.csv", safe_rows)

        session_info = {
            "tenant_id": extractor.tenant_id,
            "signup_email": extractor.signup_email,
            "batch_count": len(batches),
            "generated_at_utc": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        }
        (OUTPUT_ROOT / "tenant_session.json").write_text(json.dumps(session_info, indent=2), encoding="utf-8")
        (OUTPUT_ROOT / "feature_catalog.json").write_text(json.dumps(feature_catalog(), indent=2), encoding="utf-8")

        contract_count = sum(1 for state in states.values() if state.batch_contract)
        max_delta = max(Decimal(row["system_label_abs_delta_amount_minor"]) for row in full_rows)
        summary = {
            "tenant_id": extractor.tenant_id,
            "batch_count": len(batches),
            "batch_contract_rows_available": contract_count,
            "max_system_label_abs_delta_amount_minor": format(max_delta, "f"),
            "outputs": {
                "training_dataset_full": str((OUTPUT_ROOT / "training_dataset_full.csv").relative_to(ROOT)),
                "training_dataset_intent_safe": str((OUTPUT_ROOT / "training_dataset_intent_safe.csv").relative_to(ROOT)),
                "feature_catalog": str((OUTPUT_ROOT / "feature_catalog.json").relative_to(ROOT)),
                "tenant_session": str((OUTPUT_ROOT / "tenant_session.json").relative_to(ROOT)),
            },
        }
        (OUTPUT_ROOT / "extraction_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        print(json.dumps(summary, indent=2))
    finally:
        extractor.close()


if __name__ == "__main__":
    main()
