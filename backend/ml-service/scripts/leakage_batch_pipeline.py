from __future__ import annotations

import argparse
import csv
import json
import math
import random
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


INTENT_HEADERS = [
    "Tenant_id",
    "source_system",
    "client_batch_ref",
    "client_payout_ref",
    "invoice_id",
    "voucher_id",
    "ledger_name",
    "vendor_id",
    "vendor_name",
    "beneficiary_name",
    "beneficiary_account_number",
    "beneficiary_ifsc",
    "beneficiary_vpa",
    "amount",
    "currency",
    "payment_method",
    "rail_hint",
    "payout_purpose",
    "scheduled_execution_at",
    "expected_value_date",
    "bank_account_ref",
    "approval_ref",
    "idempotency_key",
    "remarks",
    "pan_number",
    "mcc_code",
]

SETTLEMENT_HEADERS = [
    "transaction_entity",
    "entity_id",
    "amount",
    "currency",
    "fee (exclusive tax)",
    "tax",
    "debit",
    "credit",
    "payment_method",
    "card_type",
    "issuer_name",
    "entity_created_at",
    "payment_captured_at",
    "payment_notes",
    "refund_notes",
    "arn",
    "entity_description",
    "order_id",
    "order_receipt",
    "order_notes",
    "dispute_id",
    "dispute_created_at",
    "dispute_reason",
    "settlement_id",
    "settled_at",
    "settlement_utr",
    "settled_by",
]

TWOPLACES = Decimal("0.01")


@dataclass(frozen=True)
class Scenario:
    name: str
    family: str
    has_settlement: bool
    settled_multiplier: Decimal
    settlement_delay_days: int
    blank_provider_ref: bool = False
    blank_bank_ref: bool = False
    blank_client_ref: bool = False


SCENARIOS: dict[str, Scenario] = {
    "clean": Scenario("clean", "clean", True, Decimal("1.00"), 0),
    "clean_delay_2d": Scenario("clean_delay_2d", "clean_delay", True, Decimal("1.00"), 2),
    "clean_delay_5d": Scenario("clean_delay_5d", "clean_delay", True, Decimal("1.00"), 5),
    "ref_stress_clean": Scenario(
        "ref_stress_clean",
        "ref_stress",
        True,
        Decimal("1.00"),
        1,
        blank_provider_ref=True,
        blank_bank_ref=True,
    ),
    "unmatched_intent": Scenario("unmatched_intent", "unmatched", False, Decimal("0.00"), 0),
    "under_settle_08": Scenario("under_settle_08", "under", True, Decimal("0.92"), 1),
    "under_settle_15": Scenario("under_settle_15", "under", True, Decimal("0.85"), 2),
    "under_settle_30": Scenario("under_settle_30", "under", True, Decimal("0.70"), 3),
    "under_settle_15_ref_stress": Scenario(
        "under_settle_15_ref_stress",
        "under_ref_stress",
        True,
        Decimal("0.85"),
        4,
        blank_provider_ref=True,
    ),
}

SCENARIO_WEIGHTS = {
    "clean": 0.33,
    "clean_delay_2d": 0.12,
    "clean_delay_5d": 0.05,
    "ref_stress_clean": 0.08,
    "unmatched_intent": 0.18,
    "under_settle_08": 0.10,
    "under_settle_15": 0.08,
    "under_settle_30": 0.04,
    "under_settle_15_ref_stress": 0.02,
}

BATCH_TEMPLATES = [
    {
        "template": "control_clean",
        "count": 4,
        "size_range": (190, 240),
        "scenario_mix": {"clean": 0.60, "clean_delay": 0.25, "ref_stress": 0.15},
    },
    {
        "template": "unmatched_heavy",
        "count": 4,
        "size_range": (200, 260),
        "scenario_mix": {"clean": 0.45, "clean_delay": 0.10, "unmatched": 0.35, "ref_stress": 0.10},
    },
    {
        "template": "under_heavy",
        "count": 4,
        "size_range": (200, 260),
        "scenario_mix": {"clean": 0.45, "clean_delay": 0.10, "under": 0.35, "ref_stress": 0.10},
    },
    {
        "template": "mixed_leakage",
        "count": 4,
        "size_range": (210, 280),
        "scenario_mix": {
            "clean": 0.35,
            "clean_delay": 0.10,
            "unmatched": 0.20,
            "under": 0.25,
            "ref_stress": 0.10,
        },
    },
    {
        "template": "reference_stress",
        "count": 2,
        "size_range": (180, 220),
        "scenario_mix": {"clean": 0.25, "clean_delay": 0.20, "under_ref_stress": 0.25, "ref_stress": 0.30},
    },
]


def q2(value: Decimal) -> Decimal:
    return value.quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def parse_decimal(value: Any) -> Decimal:
    text = "" if value is None else str(value).strip()
    if not text:
        return Decimal("0.00")
    return q2(Decimal(text))


def format_decimal(value: Decimal) -> str:
    return format(q2(value), "f")


def parse_dt(value: str) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]
    for pattern in candidates:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def iso_ts(day: date, source: datetime | None, fallback_hour: int, fallback_minute: int) -> str:
    src_time = source.time() if source else time(hour=fallback_hour, minute=fallback_minute)
    return datetime.combine(day, src_time).strftime("%Y-%m-%d %H:%M:%S")


def date_only(day: date) -> str:
    return day.strftime("%Y-%m-%d")


def load_intents(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for row in reader:
            rows.append({header: row.get(header, "") for header in INTENT_HEADERS})
    return rows


def load_settlements(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers = ["" if value is None else str(value).strip() for value in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        row = {}
        for header, value in zip(headers, values):
            row[header] = "" if value is None else str(value)
        rows.append({header: row.get(header, "") for header in SETTLEMENT_HEADERS})
    workbook.close()
    return rows


def build_paired_records(intents: list[dict[str, str]], settlements: list[dict[str, str]]) -> list[dict[str, Any]]:
    settlement_by_ref = {row["order_receipt"]: row for row in settlements}
    paired: list[dict[str, Any]] = []
    for index, intent in enumerate(intents, start=1):
        key = intent["client_payout_ref"]
        settlement = settlement_by_ref.get(key)
        if settlement is None:
            raise ValueError(f"missing settlement for client_payout_ref={key}")
        amount = parse_decimal(intent["amount"])
        paired.append(
            {
                "source_row_id": index,
                "rail": intent["rail_hint"] or intent["payment_method"],
                "payment_method": intent["payment_method"],
                "amount": amount,
                "currency": intent["currency"],
                "intent": intent,
                "settlement": settlement,
                "scenario": None,
            }
        )
    return paired


def interleaved_order(records: list[dict[str, Any]], rng: random.Random) -> list[dict[str, Any]]:
    amounts = sorted(record["amount"] for record in records)
    q25 = amounts[int(len(amounts) * 0.25)]
    q50 = amounts[int(len(amounts) * 0.50)]
    q75 = amounts[int(len(amounts) * 0.75)]

    def bucket_for(record: dict[str, Any]) -> tuple[str, str]:
        amount = record["amount"]
        if amount <= q25:
            size_bucket = "q1"
        elif amount <= q50:
            size_bucket = "q2"
        elif amount <= q75:
            size_bucket = "q3"
        else:
            size_bucket = "q4"
        return record["rail"], size_bucket

    buckets: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        buckets[bucket_for(record)].append(record)

    for values in buckets.values():
        rng.shuffle(values)

    ordered: list[dict[str, Any]] = []
    while True:
        progressed = False
        for key in sorted(buckets):
            if buckets[key]:
                ordered.append(buckets[key].pop())
                progressed = True
        if not progressed:
            break
    return ordered


def assign_scenarios(records: list[dict[str, Any]], seed: int) -> None:
    rng = random.Random(seed)
    ordered = interleaved_order(records, rng)
    total = len(ordered)
    assigned = 0
    scenario_counts: dict[str, int] = {}
    names = list(SCENARIO_WEIGHTS)
    cumulative = 0
    for idx, name in enumerate(names):
        if idx == len(names) - 1:
            count = total - assigned
        else:
            count = int(round(total * SCENARIO_WEIGHTS[name]))
            if assigned + count > total:
                count = total - assigned
        scenario_counts[name] = count
        assigned += count
        cumulative += count
    if cumulative != total:
        scenario_counts[names[-1]] += total - cumulative

    cursor = 0
    for name in names:
        for record in ordered[cursor : cursor + scenario_counts[name]]:
            record["scenario"] = SCENARIOS[name]
        cursor += scenario_counts[name]


def apply_scenario_truth(record: dict[str, Any]) -> dict[str, Any]:
    scenario: Scenario = record["scenario"]
    intent_amount: Decimal = record["amount"]
    expected_settlement_amount = q2(intent_amount * scenario.settled_multiplier)
    unmatched_amount = Decimal("0.00")
    under_amount = Decimal("0.00")
    if not scenario.has_settlement:
        unmatched_amount = intent_amount
        expected_settlement_amount = Decimal("0.00")
    elif expected_settlement_amount < intent_amount:
        under_amount = q2(intent_amount - expected_settlement_amount)

    total_leakage = unmatched_amount + under_amount
    return {
        **record,
        "expected_settlement_amount": expected_settlement_amount,
        "expected_unmatched_amount": unmatched_amount,
        "expected_under_settlement_amount": under_amount,
        "expected_reversal_amount": Decimal("0.00"),
        "expected_total_leakage_amount": total_leakage,
        "has_settlement": scenario.has_settlement,
    }


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]], sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def build_intent_row(
    record: dict[str, Any],
    payout_ref: str,
    batch_id: str,
    sequence_id: str,
    anchor_day: date,
    rng: random.Random,
) -> dict[str, str]:
    base = dict(record["intent"])
    scheduled_source = parse_dt(base.get("scheduled_execution_at", ""))
    base["client_batch_ref"] = batch_id
    base["client_payout_ref"] = payout_ref
    base["idempotency_key"] = f"{sequence_id}-idemp"
    base["approval_ref"] = f"{sequence_id}-approval"
    base["bank_account_ref"] = f"{sequence_id}-bank"
    base["scheduled_execution_at"] = iso_ts(anchor_day, scheduled_source, rng.randint(8, 18), rng.choice([0, 10, 20, 30, 40, 50]))
    base["expected_value_date"] = date_only(anchor_day)
    base["remarks"] = f"Generated leakage training batch {batch_id}"
    return {header: base.get(header, "") for header in INTENT_HEADERS}


def build_settlement_rows(
    record: dict[str, Any],
    payout_ref: str,
    batch_id: str,
    sequence_id: str,
    anchor_day: date,
    use_replay_credit: bool,
) -> list[dict[str, str]]:
    if not record["has_settlement"]:
        return []

    scenario: Scenario = record["scenario"]
    base = dict(record["settlement"])
    settled_amount: Decimal = record["expected_settlement_amount"]
    settlement_day = anchor_day + timedelta(days=scenario.settlement_delay_days)
    created_source = parse_dt(base.get("entity_created_at", ""))
    captured_source = parse_dt(base.get("payment_captured_at", ""))

    entity_id = "" if scenario.blank_provider_ref else f"{sequence_id}-entity"
    order_receipt = "" if scenario.blank_client_ref else payout_ref
    settlement_utr = "" if scenario.blank_bank_ref else f"{sequence_id}-utr"
    order_id = f"{sequence_id}-order"
    settlement_id = f"{batch_id}-settlement"

    row = dict(base)
    row["entity_id"] = entity_id
    row["amount"] = format_decimal(settled_amount)
    row["debit"] = format_decimal(settled_amount)
    row["credit"] = format_decimal(settled_amount if use_replay_credit else Decimal("0.00"))
    row["payment_method"] = record["payment_method"]
    row["entity_created_at"] = iso_ts(anchor_day, created_source, 9, 15)
    row["payment_captured_at"] = iso_ts(anchor_day, captured_source, 11, 30)
    row["payment_notes"] = f"batch={batch_id};source_row={record['source_row_id']};scenario={scenario.name}"
    row["entity_description"] = f"Generated payout for {batch_id}"
    row["order_id"] = order_id
    row["order_receipt"] = order_receipt
    row["order_notes"] = scenario.name
    row["settlement_id"] = settlement_id
    row["settled_at"] = date_only(settlement_day)
    row["settlement_utr"] = settlement_utr
    row["settled_by"] = "Razorpay"
    return [{header: row.get(header, "") for header in SETTLEMENT_HEADERS}]


def build_master_files(master_records: list[dict[str, Any]], output_root: Path) -> None:
    rng = random.Random(99)
    master_anchor = date(2026, 5, 1)
    intent_rows: list[dict[str, str]] = []
    settlement_truth_rows: list[dict[str, str]] = []
    settlement_replay_rows: list[dict[str, str]] = []
    manifest_rows: list[dict[str, Any]] = []

    for record in master_records:
        stable_ref = f"MASTER_PAY_{record['source_row_id']:06d}"
        batch_id = "MASTER_CORPUS"
        sequence_id = f"master-{record['source_row_id']:06d}"
        anchor_day = master_anchor + timedelta(days=(record["source_row_id"] - 1) % 28)
        intent_rows.append(build_intent_row(record, stable_ref, batch_id, sequence_id, anchor_day, rng))
        settlement_truth_rows.extend(build_settlement_rows(record, stable_ref, batch_id, sequence_id, anchor_day, False))
        settlement_replay_rows.extend(build_settlement_rows(record, stable_ref, batch_id, sequence_id, anchor_day, True))
        manifest_rows.append(
            {
                "source_row_id": record["source_row_id"],
                "scenario": record["scenario"].name,
                "scenario_family": record["scenario"].family,
                "rail": record["rail"],
                "payment_method": record["payment_method"],
                "intent_amount": format_decimal(record["amount"]),
                "expected_settlement_amount": format_decimal(record["expected_settlement_amount"]),
                "expected_unmatched_amount": format_decimal(record["expected_unmatched_amount"]),
                "expected_under_settlement_amount": format_decimal(record["expected_under_settlement_amount"]),
                "expected_reversal_amount": format_decimal(record["expected_reversal_amount"]),
                "expected_total_leakage_amount": format_decimal(record["expected_total_leakage_amount"]),
                "has_settlement": str(record["has_settlement"]).lower(),
                "settlement_delay_days": record["scenario"].settlement_delay_days,
                "blank_provider_ref": str(record["scenario"].blank_provider_ref).lower(),
                "blank_bank_ref": str(record["scenario"].blank_bank_ref).lower(),
                "blank_client_ref": str(record["scenario"].blank_client_ref).lower(),
            }
        )

    write_csv(output_root / "master_intent.csv", INTENT_HEADERS, intent_rows)
    write_xlsx(output_root / "master_settlement_truth.xlsx", SETTLEMENT_HEADERS, settlement_truth_rows, "Settlement_Mapped_2k")
    write_xlsx(output_root / "master_settlement_replay.xlsx", SETTLEMENT_HEADERS, settlement_replay_rows, "Settlement_Mapped_2k")
    write_csv(output_root / "master_truth_manifest.csv", list(manifest_rows[0].keys()), manifest_rows)


def pick_records_for_batch(
    master_records: list[dict[str, Any]],
    scenario_pools: dict[str, list[dict[str, Any]]],
    scenario_usage: Counter,
    template: dict[str, Any],
    size: int,
    rng: random.Random,
    max_reuse: int,
) -> list[dict[str, Any]]:
    scenario_targets: dict[str, int] = {}
    mix = template["scenario_mix"]
    names = list(mix)
    assigned = 0
    for index, family in enumerate(names):
        if index == len(names) - 1:
            count = size - assigned
        else:
            count = int(round(size * mix[family]))
            if assigned + count > size:
                count = size - assigned
        scenario_targets[family] = count
        assigned += count

    chosen: list[dict[str, Any]] = []
    chosen_ids: set[int] = set()
    for family, count in scenario_targets.items():
        pool = scenario_pools[family][:]
        rng.shuffle(pool)
        family_selected = 0
        for record in pool:
            if family_selected >= count:
                break
            source_row_id = record["source_row_id"]
            if source_row_id in chosen_ids:
                continue
            if scenario_usage[source_row_id] >= max_reuse:
                continue
            chosen.append(record)
            chosen_ids.add(source_row_id)
            scenario_usage[source_row_id] += 1
            family_selected += 1

    if len(chosen) < size:
        fallback = master_records[:]
        rng.shuffle(fallback)
        for record in fallback:
            if len(chosen) >= size:
                break
            source_row_id = record["source_row_id"]
            if source_row_id in chosen_ids:
                continue
            if scenario_usage[source_row_id] >= max_reuse:
                continue
            chosen.append(record)
            chosen_ids.add(source_row_id)
            scenario_usage[source_row_id] += 1

    if len(chosen) != size:
        raise RuntimeError(f"unable to materialize batch of size {size}; got {len(chosen)}")
    return chosen


def summarize_batch(batch_records: list[dict[str, Any]]) -> dict[str, Any]:
    intended = sum((record["amount"] for record in batch_records), Decimal("0.00"))
    unmatched = sum((record["expected_unmatched_amount"] for record in batch_records), Decimal("0.00"))
    under = sum((record["expected_under_settlement_amount"] for record in batch_records), Decimal("0.00"))
    reversal = sum((record["expected_reversal_amount"] for record in batch_records), Decimal("0.00"))
    leakage = unmatched + under + reversal
    leakage_rate = Decimal("0.00") if intended == 0 else q2(leakage / intended)
    family_counts = Counter(record["scenario"].family for record in batch_records)
    return {
        "batch_total_intended_amount": q2(intended),
        "expected_unmatched_amount": q2(unmatched),
        "expected_under_settlement_amount": q2(under),
        "expected_reversal_amount": q2(reversal),
        "expected_total_leakage_amount": q2(leakage),
        "expected_leakage_rate": leakage_rate,
        "scenario_family_counts": dict(family_counts),
    }


def materialize_batches(master_records: list[dict[str, Any]], output_root: Path, seed: int, max_reuse: int) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    scenario_pools: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in master_records:
        scenario_pools[record["scenario"].family].append(record)
        if record["scenario"].family.startswith("under"):
            scenario_pools["under"].append(record)
        if record["scenario"].family.startswith("clean"):
            scenario_pools["clean_delay" if "delay" in record["scenario"].family else "clean"].append(record)

    if "clean_delay" not in scenario_pools:
        scenario_pools["clean_delay"] = []
    if "clean" not in scenario_pools:
        scenario_pools["clean"] = []
    if "ref_stress" not in scenario_pools:
        scenario_pools["ref_stress"] = []
    if "under_ref_stress" not in scenario_pools:
        scenario_pools["under_ref_stress"] = []
    if "unmatched" not in scenario_pools:
        scenario_pools["unmatched"] = []
    if "under" not in scenario_pools:
        scenario_pools["under"] = []

    scenario_usage: Counter = Counter()
    batch_output: list[dict[str, Any]] = []
    anchor_start = date(2026, 5, 4)
    batch_number = 1

    for template in BATCH_TEMPLATES:
        for template_index in range(template["count"]):
            size = rng.randint(template["size_range"][0], template["size_range"][1])
            batch_records = pick_records_for_batch(
                master_records,
                scenario_pools,
                scenario_usage,
                template,
                size,
                rng,
                max_reuse,
            )
            batch_id = f"LEAK_BATCH_{batch_number:03d}"
            batch_anchor = anchor_start + timedelta(days=batch_number - 1)
            batch_dir = output_root / "batches" / batch_id
            intent_rows: list[dict[str, str]] = []
            settlement_truth_rows: list[dict[str, str]] = []
            settlement_replay_rows: list[dict[str, str]] = []
            manifest_rows: list[dict[str, Any]] = []

            for row_index, record in enumerate(batch_records, start=1):
                payout_ref = f"{batch_id}_PAY_{row_index:04d}"
                sequence_id = f"{batch_id.lower()}-{row_index:04d}"
                intent_rows.append(build_intent_row(record, payout_ref, batch_id, sequence_id, batch_anchor, rng))
                settlement_truth_rows.extend(build_settlement_rows(record, payout_ref, batch_id, sequence_id, batch_anchor, False))
                settlement_replay_rows.extend(build_settlement_rows(record, payout_ref, batch_id, sequence_id, batch_anchor, True))
                manifest_rows.append(
                    {
                        "batch_id": batch_id,
                        "batch_template": template["template"],
                        "source_row_id": record["source_row_id"],
                        "scenario": record["scenario"].name,
                        "scenario_family": record["scenario"].family,
                        "materialized_payout_ref": payout_ref,
                        "intent_amount": format_decimal(record["amount"]),
                        "expected_settlement_amount": format_decimal(record["expected_settlement_amount"]),
                        "expected_unmatched_amount": format_decimal(record["expected_unmatched_amount"]),
                        "expected_under_settlement_amount": format_decimal(record["expected_under_settlement_amount"]),
                        "expected_reversal_amount": format_decimal(record["expected_reversal_amount"]),
                        "expected_total_leakage_amount": format_decimal(record["expected_total_leakage_amount"]),
                        "has_settlement": str(record["has_settlement"]).lower(),
                        "settlement_delay_days": record["scenario"].settlement_delay_days,
                    }
                )

            summary = summarize_batch(batch_records)
            manifest = {
                "batch_id": batch_id,
                "batch_template": template["template"],
                "batch_anchor_date": batch_anchor.isoformat(),
                "batch_size": len(batch_records),
                "batch_total_intended_amount": format_decimal(summary["batch_total_intended_amount"]),
                "expected_unmatched_amount": format_decimal(summary["expected_unmatched_amount"]),
                "expected_under_settlement_amount": format_decimal(summary["expected_under_settlement_amount"]),
                "expected_reversal_amount": format_decimal(summary["expected_reversal_amount"]),
                "expected_total_leakage_amount": format_decimal(summary["expected_total_leakage_amount"]),
                "expected_leakage_rate": str(summary["expected_leakage_rate"]),
                "scenario_family_counts": summary["scenario_family_counts"],
                "truth_label": {
                    "predicted_leakage_rate": str(summary["expected_leakage_rate"]),
                    "leakage_amount_minor_assumption": format_decimal(summary["expected_total_leakage_amount"]),
                },
            }

            write_csv(batch_dir / "intent.csv", INTENT_HEADERS, intent_rows)
            write_xlsx(batch_dir / "settlement_truth.xlsx", SETTLEMENT_HEADERS, settlement_truth_rows, "Settlement_Mapped_2k")
            write_xlsx(batch_dir / "settlement_replay.xlsx", SETTLEMENT_HEADERS, settlement_replay_rows, "Settlement_Mapped_2k")
            write_csv(batch_dir / "row_truth_manifest.csv", list(manifest_rows[0].keys()), manifest_rows)
            batch_dir.mkdir(parents=True, exist_ok=True)
            (batch_dir / "batch_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

            batch_output.append(
                {
                    "batch_id": batch_id,
                    "batch_template": template["template"],
                    "batch_anchor_date": batch_anchor.isoformat(),
                    "batch_size": len(batch_records),
                    "intent_file": str((batch_dir / "intent.csv").relative_to(output_root)),
                    "settlement_truth_file": str((batch_dir / "settlement_truth.xlsx").relative_to(output_root)),
                    "settlement_replay_file": str((batch_dir / "settlement_replay.xlsx").relative_to(output_root)),
                    "expected_leakage_rate": str(summary["expected_leakage_rate"]),
                    "expected_total_leakage_amount": format_decimal(summary["expected_total_leakage_amount"]),
                    "expected_unmatched_amount": format_decimal(summary["expected_unmatched_amount"]),
                    "expected_under_settlement_amount": format_decimal(summary["expected_under_settlement_amount"]),
                    "expected_reversal_amount": format_decimal(summary["expected_reversal_amount"]),
                    "scenario_family_counts_json": json.dumps(summary["scenario_family_counts"], sort_keys=True),
                }
            )
            batch_number += 1

    return batch_output


def emit_summary(output_root: Path, master_records: list[dict[str, Any]], batch_output: list[dict[str, Any]], seed: int, max_reuse: int) -> None:
    scenario_counts = Counter(record["scenario"].name for record in master_records)
    total_intended = sum((record["amount"] for record in master_records), Decimal("0.00"))
    total_leakage = sum((record["expected_total_leakage_amount"] for record in master_records), Decimal("0.00"))
    summary = {
        "seed": seed,
        "max_reuse_per_source_row": max_reuse,
        "master_record_count": len(master_records),
        "batch_count": len(batch_output),
        "master_total_intended_amount": format_decimal(total_intended),
        "master_total_expected_leakage_amount": format_decimal(total_leakage),
        "master_expected_leakage_rate": str(q2(total_leakage / total_intended) if total_intended else Decimal("0.00")),
        "scenario_counts": dict(sorted(scenario_counts.items())),
    }
    (output_root / "generation_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a leaky master corpus and replay-ready batch files.")
    parser.add_argument("--intent-file", type=Path, default=Path("backend/zord_payout_v4_final.csv"))
    parser.add_argument("--settlement-file", type=Path, default=Path("backend/Razorpay_Settlement_v4.xlsx"))
    parser.add_argument("--output-root", type=Path, default=Path("backend/generated/leakage_training"))
    parser.add_argument("--seed", type=int, default=20260611)
    parser.add_argument("--max-reuse", type=int, default=3)
    args = parser.parse_args()

    intents = load_intents(args.intent_file)
    settlements = load_settlements(args.settlement_file)
    paired = build_paired_records(intents, settlements)
    assign_scenarios(paired, args.seed)
    master_records = [apply_scenario_truth(record) for record in paired]

    args.output_root.mkdir(parents=True, exist_ok=True)
    build_master_files(master_records, args.output_root)
    batch_output = materialize_batches(master_records, args.output_root, args.seed + 1, args.max_reuse)
    write_csv(args.output_root / "batch_index.csv", list(batch_output[0].keys()), batch_output)
    emit_summary(args.output_root, master_records, batch_output, args.seed, args.max_reuse)

    print(json.dumps(
        {
            "output_root": str(args.output_root),
            "master_records": len(master_records),
            "batches_generated": len(batch_output),
            "batch_index": str(args.output_root / "batch_index.csv"),
        },
        indent=2,
    ))


if __name__ == "__main__":
    main()
